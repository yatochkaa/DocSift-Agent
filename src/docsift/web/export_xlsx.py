"""Production XLSX export for reviewed DocSift documents.

The workbook intentionally separates the printable business document from the
structured positions, review history, and guardrail audit trail.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string


DARK = "243142"
BLUE = "2E5E82"
MUTED = "5E6B76"
LINE = "AAB7C2"
SOFT = "EAF1F5"
PALE = "F6F8FA"
SUCCESS_BG = "FFE8F5E9"
SUCCESS_TEXT = "1B5E3A"
ERROR_BG = "FFFEE2E2"
ERROR_TEXT = "991B1B"
ATTENTION_BG = "FFFFF3CD"
WHITE = "FFFFFF"

THIN = Side(style="thin", color=LINE)
MEDIUM = Side(style="medium", color=DARK)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FIELD_LABELS = {
    "/date": "Дата документа",
    "/buyer/inn": "ИНН покупателя",
    "/buyer/kpp": "КПП покупателя",
    "/buyer/name": "Покупатель",
    "/number": "Номер",
    "/currency": "Валюта",
    "/supplier/inn": "ИНН поставщика",
    "/supplier/kpp": "КПП поставщика",
    "/supplier/name": "Поставщик",
    "/vat_amount": "НДС",
    "/total_amount": "Итого",
    "/document_type": "Тип документа",
}

DOCUMENT_TYPE_LABELS = {
    "payment_invoice": "Счёт на оплату",
    "invoice": "Счёт",
    "act": "Акт",
    "contract": "Договор",
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, Decimal):
        return str(int(value)) if value == value.to_integral() else str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _num(value: Any) -> float | None:
    """Return a number or ``None`` for absent/invalid optional values."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    return None


def _num_or_zero(value: Any) -> float:
    parsed = _num(value)
    return parsed if parsed is not None else 0.0


def _item_value(item: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in item:
            return item.get(key)
    return None



def _safe_cell(ws: Any, row: int, col: int, value: str) -> None:
    """Write a string cell with XLSX-native formula-injection protection.

    Sets ``data_type = "s"`` (inline string) so openpyxl writes ``<c t="inlineStr">``
    in the XML, preventing formula interpretation by Excel/LibreOffice.  The cell
    value is **not** modified — no leading apostrophe is prepended.
    """
    cell = ws.cell(row, col, value)
    cell.data_type = "s"


def _field_map(payload: Mapping[str, Any]) -> dict[str, Mapping[str, Any]]:
    fields = payload.get("extracted", {}).get("fields", []) or []
    return {
        str(field.get("path")): field
        for field in fields
        if isinstance(field, Mapping) and field.get("path")
    }


def _field_value(fields: Mapping[str, Mapping[str, Any]], path: str) -> Any:
    field = fields.get(path)
    return field.get("value") if field else None


def _format_russian_date(value: Any) -> str:
    if value is None:
        return ""
    parsed = value
    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00")).date()
        except ValueError:
            return value
    if isinstance(parsed, datetime):
        parsed = parsed.date()
    if isinstance(parsed, date):
        months = (
            "января", "февраля", "марта", "апреля", "мая", "июня",
            "июля", "августа", "сентября", "октября", "ноября", "декабря",
        )
        return f"«{parsed.day}» {months[parsed.month - 1]} {parsed.year} г."
    return _text(value)


def _integer_words(number: int) -> str:
    """Russian cardinal words for non-negative amounts up to billions."""
    if number == 0:
        return "ноль"
    ones_m = ("", "один", "два", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять")
    ones_f = ("", "одна", "две", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять")
    teens = ("десять", "одиннадцать", "двенадцать", "тринадцать", "четырнадцать", "пятнадцать", "шестнадцать", "семнадцать", "восемнадцать", "девятнадцать")
    tens = ("", "", "двадцать", "тридцать", "сорок", "пятьдесят", "шестьдесят", "семьдесят", "восемьдесят", "девяносто")
    hundreds = ("", "сто", "двести", "триста", "четыреста", "пятьсот", "шестьсот", "семьсот", "восемьсот", "девятьсот")
    groups = (
        ("", "", "", False),
        ("тысяча", "тысячи", "тысяч", True),
        ("миллион", "миллиона", "миллионов", False),
        ("миллиард", "миллиарда", "миллиардов", False),
    )

    def form(value: int, forms: tuple[str, str, str]) -> str:
        last_two = value % 100
        last = value % 10
        if 11 <= last_two <= 19:
            return forms[2]
        if last == 1:
            return forms[0]
        if 2 <= last <= 4:
            return forms[1]
        return forms[2]

    parts: list[str] = []
    group_index = 0
    while number:
        value = number % 1000
        if value:
            chunk: list[str] = []
            chunk.append(hundreds[value // 100])
            remainder = value % 100
            if 10 <= remainder <= 19:
                chunk.append(teens[remainder - 10])
            else:
                chunk.append(tens[remainder // 10])
                ones = ones_f if groups[group_index][3] else ones_m
                chunk.append(ones[remainder % 10])
            if group_index:
                chunk.append(form(value, groups[group_index][:3]))
            parts = [word for word in chunk if word] + parts
        number //= 1000
        group_index += 1
    return " ".join(parts)


def _money_in_words(value: float) -> str:
    value = round(value, 2)
    rubles = int(abs(value))
    kopeks = int(round((abs(value) - rubles) * 100))
    sign = "минус " if value < 0 else ""
    words = _integer_words(rubles)
    return f"{sign}{words.capitalize()} рублей {kopeks:02d} копеек"


def _set_print(
    ws: Any,
    print_area: str,
    *,
    orientation: str,
    fit_height: int = 0,
    compact: bool = False,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = print_area
    ws.page_margins.left = 0.28 if compact else 0.3
    ws.page_margins.right = 0.28 if compact else 0.3
    ws.page_margins.top = 0.35 if compact else 0.5
    ws.page_margins.bottom = 0.4 if compact else 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.3
    ws.oddFooter.right.text = "Страница &P из &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = MUTED


def _style_header(cell: Any) -> None:
    cell.font = Font(name="Arial", size=9, bold=True, color=DARK)
    cell.fill = PatternFill("solid", fgColor=SOFT)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(top=MEDIUM, bottom=THIN)


def _party_text(fields: Mapping[str, Mapping[str, Any]], prefix: str) -> tuple[str, str]:
    name = _text(_field_value(fields, f"/{prefix}/name"))
    details = []
    inn = _field_value(fields, f"/{prefix}/inn")
    kpp = _field_value(fields, f"/{prefix}/kpp")
    if inn not in (None, ""):
        details.append(f"ИНН {_text(inn)}")
    if kpp not in (None, ""):
        details.append(f"КПП {_text(kpp)}")
    return name, "    ".join(details)


def _build_document_sheet(ws: Any, payload: Mapping[str, Any]) -> None:
    ws.title = "Документ"
    fields = _field_map(payload)
    document = payload.get("document", {}) or {}
    line_items = list(payload.get("extracted", {}).get("line_items", []) or [])
    guardrails = list(payload.get("guardrails", []) or [])

    document_type = _field_value(fields, "/document_type") or document.get("document_type")
    type_label = DOCUMENT_TYPE_LABELS.get(_text(document_type), "Счёт на оплату")
    number = _text(_field_value(fields, "/number"))
    date_text = _format_russian_date(_field_value(fields, "/date"))
    title = type_label
    if number:
        title += f" № {number}"
    if date_text:
        title += f" от {date_text}"

    file_name = _text(document.get("file_name") or document.get("source_name") or document.get("id"))
    corrected_count = sum(1 for field in fields.values() if field.get("corrected"))
    all_guardrails_passed = all(rule.get("passed") for rule in guardrails) if guardrails else True

    widths = {"A": 4, "B": 7, "C": 29, "D": 9, "E": 10, "F": 12, "G": 13, "H": 11, "I": 13, "J": 18}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    _safe_cell(ws, 1, 1, "DocSift  /  ПРОВЕРЕННЫЙ ДОКУМЕНТ")
    ws["A1"].font = Font(name="Arial", size=9, bold=True, color=BLUE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A1:F1")
    _safe_cell(ws, 1, 7, "ПРОВЕРЕНО")
    ws["G1"].font = Font(name="Arial", size=10, bold=True, color="237A45")
    ws["G1"].fill = PatternFill("solid", fgColor="E9F5EE")
    ws["G1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["G1"].border = CELL_BORDER
    ws.merge_cells("G1:J1")
    ws.row_dimensions[1].height = 21.75

    _safe_cell(ws, 3, 1, title)
    ws["A3"].font = Font(name="Arial", size=17, bold=True, color=DARK)
    ws["A3"].alignment = Alignment(vertical="center")
    ws["A3"].border = Border(bottom=MEDIUM)
    ws.merge_cells("A3:J3")
    ws.row_dimensions[3].height = 33.75

    source_text = f"Источник: {file_name}" if file_name else "Источник: не указан"
    _safe_cell(ws, 4, 1, source_text)
    ws["A4"].font = Font(name="Arial", size=8, color=MUTED)
    ws.merge_cells("A4:F4")
    _safe_cell(ws, 4, 7, f"Исправлено полей: {corrected_count}")
    ws["G4"].font = Font(name="Arial", size=8, color=MUTED)
    ws["G4"].alignment = Alignment(horizontal="right")
    ws.merge_cells("G4:J4")

    supplier_name, supplier_details = _party_text(fields, "supplier")
    buyer_name, buyer_details = _party_text(fields, "buyer")
    for start, end, heading, name, details in (
        ("A", "E", "ПОСТАВЩИК", supplier_name, supplier_details),
        ("F", "J", "ПОКУПАТЕЛЬ", buyer_name, buyer_details),
    ):
        heading_cell = ws[f"{start}6"]
        heading_cell.value = heading
        heading_cell.font = Font(name="Arial", size=9, bold=True, color=BLUE)
        heading_cell.fill = PatternFill("solid", fgColor=SOFT)
        heading_cell.border = CELL_BORDER
        ws.merge_cells(f"{start}6:{end}6")
        _safe_cell(ws, 7, column_index_from_string(start), name or "Не указан")
        name_cell = ws[f"{start}7"]
        name_cell.font = Font(name="Arial", size=10, bold=True, color=DARK)
        name_cell.alignment = Alignment(vertical="center", wrap_text=True)
        name_cell.border = Border(left=THIN, right=THIN)
        ws.merge_cells(f"{start}7:{end}8")
        _safe_cell(ws, 9, column_index_from_string(start), details)
        details_cell = ws[f"{start}9"]
        details_cell.font = Font(name="Arial", size=8, color=MUTED)
        details_cell.alignment = Alignment(vertical="center")
        details_cell.border = Border(left=THIN, right=THIN, bottom=THIN)
        ws.merge_cells(f"{start}9:{end}9")
    ws.row_dimensions[7].height = 24
    ws.row_dimensions[8].height = 24

    header_row = 12
    headers = (
        "№", "Наименование товара, работы, услуги", None, "Ед.", "Кол-во",
        "Цена", "Сумма без НДС", "Ставка НДС", "Сумма НДС", "Всего с НДС",
    )
    for col, value in enumerate(headers, 1):
        if value is not None:
            ws.cell(header_row, col, value)
        _style_header(ws.cell(header_row, col))
    ws.merge_cells("B12:C12")
    ws.row_dimensions[12].height = 31.5

    data_start = 13
    if line_items:
        for offset, item in enumerate(line_items):
            row = data_start + offset
            quantity = _num(_item_value(item, "quantity", "qty"))
            price = _num(_item_value(item, "unit_price", "price"))
            amount = _num(_item_value(item, "amount", "subtotal"))
            vat_amount = _num(_item_value(item, "vat_amount"))
            total = _num(_item_value(item, "total", "total_amount"))
            vat_rate = _item_value(item, "vat_rate")
            if total is None and (amount is not None or vat_amount is not None):
                total = _num_or_zero(amount) + _num_or_zero(vat_amount)
            values = (
                offset + 1,
                _text(item.get("name")),
                None,
                _text(_item_value(item, "unit", "unit_name")),
                quantity,
                price,
                amount,
                vat_rate,
                vat_amount,
                total,
            )
            for col, value in enumerate(values, 1):
                if value is not None:
                    if isinstance(value, str):
                        _safe_cell(ws, row, col, value)
                    else:
                        ws.cell(row, col, value)
                cell = ws.cell(row, col)
                cell.font = Font(name="Arial", size=8, color=DARK)
                cell.alignment = Alignment(
                    horizontal="center" if col == 1 else ("left" if col in (2, 3) else "right"),
                    vertical="center",
                    wrap_text=col in (2, 3),
                    indent=1 if col == 2 else 0,
                )
                cell.border = Border(bottom=THIN)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws.cell(row, 1).number_format = "0"
            ws.cell(row, 5).number_format = "#,##0.###"
            for col in (6, 7, 9, 10):
                ws.cell(row, col).number_format = "#,##0.00"
            if vat_rate == "without_vat":
                _safe_cell(ws, row, 8, "без НДС")
            elif _num(vat_rate) is not None:
                ws.cell(row, 8).number_format = "0.##%"
                numeric_rate = _num(vat_rate)
                ws.cell(row, 8, numeric_rate / 100 if numeric_rate and numeric_rate > 1 else numeric_rate)
            ws.row_dimensions[row].height = 33
        data_end = data_start + len(line_items) - 1
        ws.auto_filter.ref = f"A12:J{data_end}"
        ws.freeze_panes = "A13"
    else:
        data_end = data_start
        _safe_cell(ws, 13, 1, "В документе нет табличных позиций")
        ws["A13"].font = Font(name="Arial", size=9, italic=True, color=MUTED)
        ws["A13"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A13"].border = Border(bottom=THIN)
        ws.merge_cells("A13:J13")
        ws.row_dimensions[13].height = 30

    subtotal = sum(_num_or_zero(_item_value(item, "amount", "subtotal")) for item in line_items)
    total_vat = sum(_num_or_zero(_item_value(item, "vat_amount")) for item in line_items)
    total_from_items = sum(
        _num_or_zero(_item_value(item, "total", "total_amount"))
        if _num(_item_value(item, "total", "total_amount")) is not None
        else _num_or_zero(_item_value(item, "amount", "subtotal")) + _num_or_zero(_item_value(item, "vat_amount"))
        for item in line_items
    )
    payload_total = _num(_field_value(fields, "/total_amount"))
    payload_vat = _num(_field_value(fields, "/vat_amount"))
    grand_total = total_from_items if line_items else _num_or_zero(payload_total)
    if not line_items:
        total_vat = _num_or_zero(payload_vat)
        subtotal = grand_total - total_vat

    totals_start = data_end + 2
    for row, label, value, emphasized in (
        (totals_start, "Итого без НДС", subtotal, False),
        (totals_start + 1, "НДС", total_vat, False),
        (totals_start + 2, "Всего к оплате", grand_total, True),
    ):
        _safe_cell(ws, row, 6, label)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
        ws.cell(row, 10, value)
        ws.cell(row, 10).number_format = '#,##0.00 "₽"'
        for col in range(6, 11):
            cell = ws.cell(row, col)
            cell.font = Font(name="Arial", size=10 if emphasized else 8, bold=True, color=DARK)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if emphasized:
                cell.fill = PatternFill("solid", fgColor=SOFT)
                cell.border = Border(top=MEDIUM, bottom=MEDIUM)
        ws.row_dimensions[row].height = 27.75 if emphasized else 15

    summary_row = totals_start + 4
    count = len(line_items)
    _safe_cell(ws, summary_row, 1, f"Всего наименований: {count}, на сумму {grand_total:,.2f} руб.")
    ws.cell(summary_row, 1).font = Font(name="Arial", size=8, color=MUTED)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=10)
    words_row = summary_row + 1
    _safe_cell(ws, words_row, 1, _money_in_words(grand_total))
    ws.cell(words_row, 1).font = Font(name="Arial", size=9, bold=True, color=DARK)
    ws.cell(words_row, 1).border = Border(bottom=MEDIUM)
    ws.merge_cells(start_row=words_row, start_column=1, end_row=words_row, end_column=10)

    review_heading_row = words_row + 2
    _safe_cell(ws, review_heading_row, 1, "СВЕДЕНИЯ О ПРОВЕРКЕ")
    ws.cell(review_heading_row, 1).font = Font(name="Arial", size=9, bold=True, color=BLUE)
    ws.cell(review_heading_row, 1).fill = PatternFill("solid", fgColor=SOFT)
    ws.merge_cells(start_row=review_heading_row, start_column=1, end_row=review_heading_row, end_column=10)
    status_row = review_heading_row + 1
    checks_text = "пройдены" if all_guardrails_passed else "требуют внимания"
    _safe_cell(
        ws,
        status_row,
        1,
        f"Статус: проверка завершена   •   Ручных исправлений: {corrected_count}   •   Контрольные проверки: {checks_text}",
    )
    ws.cell(status_row, 1).font = Font(name="Arial", size=8, color=DARK)
    ws.merge_cells(start_row=status_row, start_column=1, end_row=status_row, end_column=10)
    note_row = status_row + 1
    _safe_cell(
        ws,
        note_row,
        1,
        "Исходные данные модели сохранены отдельно. Итоговые значения учитывают подтверждённые ручные исправления.",
    )
    ws.cell(note_row, 1).font = Font(name="Arial", size=8, italic=True, color=MUTED)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)

    reviewer_row = note_row + 3
    _safe_cell(ws, reviewer_row, 1, "Ответственный за проверку  __________________________")
    ws.cell(reviewer_row, 1).font = Font(name="Arial", size=8, color=DARK)
    ws.merge_cells(start_row=reviewer_row, start_column=1, end_row=reviewer_row, end_column=5)
    _safe_cell(ws, reviewer_row, 7, "Дата  __________________")
    ws.cell(reviewer_row, 7).font = Font(name="Arial", size=8, color=DARK)
    ws.merge_cells(start_row=reviewer_row, start_column=7, end_row=reviewer_row, end_column=10)

    _set_print(ws, f"A1:J{reviewer_row + 1}", orientation="portrait", fit_height=1, compact=True)
    ws.oddFooter.left.text = f"Источник: {file_name}" if file_name else "DocSift"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = MUTED


def _build_positions_sheet(ws: Any, payload: Mapping[str, Any]) -> None:
    ws.title = "Позиции"
    line_items = list(payload.get("extracted", {}).get("line_items", []) or [])
    _safe_cell(ws, 1, 1, "Позиции документа")
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color=DARK)
    ws.merge_cells("A1:I1")
    _safe_cell(ws, 2, 1, "Структурированные итоговые значения после ручной проверки")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color=MUTED)
    ws.merge_cells("A2:I2")

    widths = {"A": 5, "B": 47, "C": 10, "D": 12, "E": 14, "F": 17, "G": 13, "H": 15, "I": 17}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    if not line_items:
        _safe_cell(ws, 4, 1, "В документе нет табличных позиций")
        ws["A4"].font = Font(name="Arial", size=10, italic=True, color=MUTED)
        ws.merge_cells("A4:I4")
        _set_print(ws, "A1:I4", orientation="landscape")
        return

    headers = ("№", "Наименование", "Ед. изм.", "Количество", "Цена", "Сумма без НДС", "Ставка НДС", "Сумма НДС", "Всего с НДС")
    for col, header in enumerate(headers, 1):
        _safe_cell(ws, 4, col, header)
        _style_header(ws.cell(4, col))
    ws.row_dimensions[4].height = 27.75

    for offset, item in enumerate(line_items):
        row = 5 + offset
        amount = _num(_item_value(item, "amount", "subtotal"))
        vat_amount = _num(_item_value(item, "vat_amount"))
        total = _num(_item_value(item, "total", "total_amount"))
        vat_rate = _item_value(item, "vat_rate")
        if total is None and (amount is not None or vat_amount is not None):
            total = _num_or_zero(amount) + _num_or_zero(vat_amount)
        values = (
            offset + 1,
            _text(item.get("name")),
            _text(_item_value(item, "unit", "unit_name")),
            _num(_item_value(item, "quantity", "qty")),
            _num(_item_value(item, "unit_price", "price")),
            amount,
            vat_rate,
            vat_amount,
            total,
        )
        for col, value in enumerate(values, 1):
            if value is not None:
                if isinstance(value, str):
                    _safe_cell(ws, row, col, value)
                else:
                    ws.cell(row, col, value)
            cell = ws.cell(row, col)
            cell.font = Font(name="Arial", size=9, color=DARK)
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else ("left" if col in (2, 3) else "right"),
                vertical="center",
                wrap_text=True,
                indent=1 if col == 2 else 0,
            )
            cell.border = Border(bottom=THIN)
        ws.cell(row, 1).number_format = "0"
        ws.cell(row, 4).number_format = "#,##0.###"
        for col in (5, 6, 8, 9):
            ws.cell(row, col).number_format = "#,##0.00"
        if vat_rate == "without_vat":
            _safe_cell(ws, row, 7, "без НДС")
        elif _num(vat_rate) is not None:
            numeric_rate = _num(vat_rate)
            ws.cell(row, 7, numeric_rate / 100 if numeric_rate and numeric_rate > 1 else numeric_rate)
            ws.cell(row, 7).number_format = "0.##%"
        ws.row_dimensions[row].height = 30

    totals_row = 5 + len(line_items)
    subtotal = sum(_num_or_zero(_item_value(item, "amount", "subtotal")) for item in line_items)
    vat_total = sum(_num_or_zero(_item_value(item, "vat_amount")) for item in line_items)
    grand_total = sum(
        _num_or_zero(_item_value(item, "total", "total_amount"))
        if _num(_item_value(item, "total", "total_amount")) is not None
        else _num_or_zero(_item_value(item, "amount", "subtotal")) + _num_or_zero(_item_value(item, "vat_amount"))
        for item in line_items
    )
    _safe_cell(ws, totals_row, 1, "ИТОГО")
    ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=5)
    ws.cell(totals_row, 6, subtotal)
    ws.cell(totals_row, 8, vat_total)
    ws.cell(totals_row, 9, grand_total)
    for col in range(1, 10):
        cell = ws.cell(totals_row, col)
        cell.font = Font(name="Arial", size=10, bold=True, color=DARK)
        cell.border = Border(top=MEDIUM, bottom=MEDIUM)
        if col in (6, 8, 9):
            cell.number_format = "#,##0.00"
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{totals_row - 1}"
    _set_print(ws, f"A1:I{totals_row}", orientation="landscape")


def _build_review_sheet(ws: Any, payload: Mapping[str, Any]) -> None:
    ws.title = "Проверка"
    fields = list(payload.get("extracted", {}).get("fields", []) or [])
    _safe_cell(ws, 1, 1, "Журнал ручной проверки")
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color=DARK)
    ws.merge_cells("A1:F1")
    _safe_cell(ws, 2, 1, "Жёлтым отмечены значения, которые были изменены пользователем")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color=MUTED)
    ws.merge_cells("A2:F2")
    headers = ("Поле", "Значение модели", "Итоговое значение", "Уверенность", "Статус", "Технический путь")
    for col, header in enumerate(headers, 1):
        _safe_cell(ws, 4, col, header)
        _style_header(ws.cell(4, col))
    for offset, field in enumerate(fields):
        row = 5 + offset
        path = _text(field.get("path"))
        corrected = bool(field.get("corrected"))
        values = (
            _text(field.get("name")) or FIELD_LABELS.get(path, path),
            _text(field.get("original_value")),
            _text(field.get("value")),
            _num(field.get("confidence")),
            "Исправлено" if corrected else "Без изменений",
            path,
        )
        for col, value in enumerate(values, 1):
            if value is not None:
                if isinstance(value, str):
                    _safe_cell(ws, row, col, value)
                else:
                    ws.cell(row, col, value)
            cell = ws.cell(row, col)
            cell.font = Font(name="Arial", size=9, color=DARK)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = CELL_BORDER
            if corrected:
                cell.fill = PatternFill("solid", fgColor=ATTENTION_BG)
        ws.cell(row, 4).number_format = "0.0%"
        ws.row_dimensions[row].height = 28.5
    widths = {"A": 27, "B": 35, "C": 35, "D": 13, "E": 16, "F": 25}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    last_row = max(4, 4 + len(fields))
    ws.freeze_panes = "A5"
    if fields:
        ws.auto_filter.ref = f"A4:F{last_row}"
    _set_print(ws, f"A1:F{last_row}", orientation="landscape")


def _guardrail_path(rule: Mapping[str, Any]) -> str:
    explicit = rule.get("path")
    if explicit:
        return _text(explicit)
    rule_name = _text(rule.get("rule"))
    if rule_name.startswith("/"):
        return rule_name
    message = _text(rule.get("message"))
    first = message.split(":", 1)[0]
    return first if first.startswith("/") else ""


def _build_guardrails_sheet(ws: Any, payload: Mapping[str, Any]) -> None:
    ws.title = "Guardrails"
    _safe_cell(ws, 1, 1, "Контрольные проверки")
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color=DARK)
    ws.merge_cells("A1:D1")
    _safe_cell(ws, 2, 1, "Итоговое состояние проверок после применения ручных исправлений")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color=MUTED)
    ws.merge_cells("A2:D2")
    headers = ("Проверка", "Результат", "Пояснение", "Технический путь")
    for col, header in enumerate(headers, 1):
        _safe_cell(ws, 4, col, header)
        _style_header(ws.cell(4, col))

    seen: set[tuple[str, bool, str]] = set()
    guardrails: list[Mapping[str, Any]] = []
    for rule in payload.get("guardrails", []) or []:
        key = (_text(rule.get("rule")), bool(rule.get("passed")), _text(rule.get("message")))
        if key not in seen:
            seen.add(key)
            guardrails.append(rule)

    if not guardrails:
        _safe_cell(ws, 5, 1, "Нарушений не зарегистрировано")
        ws["A5"].font = Font(name="Arial", size=9, italic=True, color=MUTED)
        ws.merge_cells("A5:D5")
        last_row = 5
    else:
        for offset, rule in enumerate(guardrails):
            row = 5 + offset
            path = _guardrail_path(rule)
            passed = bool(rule.get("passed"))
            message = _text(rule.get("message"))
            explanation = message.split(":", 1)[1].strip() if ":" in message else message
            values = (
                FIELD_LABELS.get(path, _text(rule.get("rule")) or path),
                "Пройдено" if passed else "Требует внимания",
                explanation,
                path,
            )
            for col, value in enumerate(values, 1):
                if isinstance(value, str):
                    _safe_cell(ws, row, col, value)
                else:
                    ws.cell(row, col, value)
                cell = ws.cell(row, col)
                cell.font = Font(name="Arial", size=9, color=DARK)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = CELL_BORDER
            result_cell = ws.cell(row, 2)
            result_cell.font = Font(
                name="Arial",
                size=9,
                bold=True,
                color=SUCCESS_TEXT if passed else ERROR_TEXT,
            )
            result_cell.fill = PatternFill("solid", fgColor=SUCCESS_BG if passed else ERROR_BG)
            ws.row_dimensions[row].height = 28.5
        last_row = 4 + len(guardrails)
        ws.auto_filter.ref = f"A4:D{last_row}"
    widths = {"A": 28, "B": 16, "C": 68, "D": 24}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A5"
    _set_print(ws, f"A1:D{last_row}", orientation="landscape")


def build_document_xlsx(payload: Mapping[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _build_document_sheet(wb.create_sheet("Документ"), payload)
    _build_positions_sheet(wb.create_sheet("Позиции"), payload)
    _build_review_sheet(wb.create_sheet("Проверка"), payload)
    _build_guardrails_sheet(wb.create_sheet("Guardrails"), payload)
    wb.active = 0
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

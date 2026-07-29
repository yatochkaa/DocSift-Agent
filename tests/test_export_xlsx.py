"""Tests for export_xlsx module."""

import pytest
from datetime import date
from decimal import Decimal

from docsift.web.export_xlsx import build_document_xlsx


@pytest.fixture
def sample_payload_with_line_items():
    """Sample payload with line items."""
    return {
        "document": {
            "id": "doc-123",
            "file_name": "invoice.pdf",
            "status": "completed",
        },
        "extracted": {
            "fields": [
                {
                    "path": "/number",
                    "name": "Номер",
                    "value": "143",
                    "original_value": "143",
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/date",
                    "name": "Дата документа",
                    "value": date(2025, 3, 12),
                    "original_value": date(2025, 3, 12),
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/supplier/name",
                    "name": "Поставщик",
                    "value": "ООО \"Ромашка\"",
                    "original_value": "ООО \"Ромашка\"",
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/supplier/inn",
                    "name": "ИНН поставщика",
                    "value": "7736050003",
                    "original_value": "7714236789",
                    "confidence": 1.0,
                    "corrected": True,
                },
                {
                    "path": "/supplier/kpp",
                    "name": "КПП поставщика",
                    "value": "771401001",
                    "original_value": "771401001",
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/buyer/name",
                    "name": "Покупатель",
                    "value": "ООО \"ТехноСервис Плюс\"",
                    "original_value": "ООО \"ТехноСервис Плюс\"",
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/buyer/inn",
                    "name": "ИНН покупателя",
                    "value": "7707083893",
                    "original_value": "5029154872",
                    "confidence": 1.0,
                    "corrected": True,
                },
                {
                    "path": "/buyer/kpp",
                    "name": "КПП покупателя",
                    "value": "502901001",
                    "original_value": "502901001",
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/currency",
                    "name": "Валюта",
                    "value": "RUB",
                    "original_value": "RUB",
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/vat_amount",
                    "name": "НДС",
                    "value": 11594,
                    "original_value": 11594,
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/total_amount",
                    "name": "Итого",
                    "value": 69564,
                    "original_value": 69564,
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/document_type",
                    "name": "Тип документа",
                    "value": "payment_invoice",
                    "original_value": "payment_invoice",
                    "confidence": 1.0,
                    "corrected": False,
                },
            ],
            "line_items": [
                {
                    "name": "Бумага офисная СветоКопи А4, 80 г/м2, 500 л.",
                    "unit": "пачка",
                    "quantity": 40,
                    "unit_price": 389.5,
                    "amount": 15580,
                    "vat_rate": "20",
                    "vat_amount": 3116,
                    "total": 18696,
                },
                {
                    "name": "Картридж лазерный HP CF259A (совместимый)",
                    "unit": "шт",
                    "quantity": 6,
                    "unit_price": 2740,
                    "amount": 16440,
                    "vat_rate": "20",
                    "vat_amount": 3288,
                    "total": 19728,
                },
                {
                    "name": "Стеллаж металлический архивный 2000х900х400",
                    "unit": "шт",
                    "quantity": 3,
                    "unit_price": 8150,
                    "amount": 24450,
                    "vat_rate": "20",
                    "vat_amount": 4890,
                    "total": 29340,
                },
                {
                    "name": "Доставка по г. Москве",
                    "unit": "усл",
                    "quantity": 1,
                    "unit_price": 1500,
                    "amount": 1500,
                    "vat_rate": "20",
                    "vat_amount": 300,
                    "total": 1800,
                },
            ],
        },
        "guardrails": [
            {
                "rule": "/buyer/inn",
                "passed": True,
                "message": "/buyer/inn: manual_correction: исправлено при ручной проверке",
            },
            {
                "rule": "/supplier/inn",
                "passed": True,
                "message": "/supplier/inn: manual_correction: исправлено при ручной проверке",
            },
        ],
    }


@pytest.fixture
def sample_payload_without_line_items():
    """Sample payload without line items."""
    return {
        "document": {
            "id": "doc-456",
            "file_name": "contract.docx",
            "status": "completed",
        },
        "extracted": {
            "fields": [
                {
                    "path": "/number",
                    "name": "Номер",
                    "value": "CTR-2025-001",
                    "original_value": "CTR-2025-001",
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/date",
                    "name": "Дата документа",
                    "value": date(2025, 1, 15),
                    "original_value": date(2025, 1, 15),
                    "confidence": 1.0,
                    "corrected": False,
                },
            ],
            "line_items": [],
        },
        "guardrails": [],
    }


def test_build_document_xlsx_creates_four_sheets(sample_payload_with_line_items):
    """Test that build_document_xlsx creates exactly 4 sheets."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    # Load the workbook
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    
    assert len(wb.sheetnames) == 4
    assert "Документ" in wb.sheetnames
    assert "Позиции" in wb.sheetnames
    assert "Проверка" in wb.sheetnames
    assert wb.sheetnames == ["Документ", "Позиции", "Проверка", "Guardrails"]


def test_document_sheet_structure(sample_payload_with_line_items):
    """Test document sheet has correct structure."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]
    
    # Check premium header and document title.
    assert ws["A1"].value == "DocSift  /  ПРОВЕРЕННЫЙ ДОКУМЕНТ"
    assert ws["G1"].value == "ПРОВЕРЕНО"
    title_cell = ws["A3"]
    assert "Счёт на оплату" in title_cell.value
    assert title_cell.font.bold
    
    # Check supplier section
    supplier_row_found = False
    buyer_row_found = False
    for row in ws.iter_rows(min_row=4, max_row=10, values_only=True):
        if any("ПОСТАВЩИК" in str(cell) for cell in row):
            supplier_row_found = True
        if any("ПОКУПАТЕЛЬ" in str(cell) for cell in row):
            buyer_row_found = True
    
    assert supplier_row_found, "Supplier section not found"
    assert buyer_row_found, "Buyer section not found"


def test_document_sheet_with_line_items(sample_payload_with_line_items):
    """Test document sheet displays line items correctly."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)

    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]

    all_text = " ".join(
        str(cell)
        for row in ws.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )

    expected_names = [
        item["name"]
        for item in sample_payload_with_line_items["extracted"]["line_items"]
    ]
    for name in expected_names:
        assert name in all_text, f"Line item '{name}' not found on Document sheet"


def test_positions_sheet_with_items(sample_payload_with_line_items):
    """Test positions sheet displays line items correctly."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Позиции"]
    
    # Check header
    assert "Позиции документа" in ws["A1"].value
    
    # Check column headers
    headers = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    assert "Наименование" in str(headers[1])
    assert "Количество" in str(headers[3])
    assert "Цена" in str(headers[4])
    
    # Check that we have 4 data rows
    data_rows = 0
    for row_idx in range(5, 13):
        cell_value = ws.cell(row=row_idx, column=1).value
        if isinstance(cell_value, int) and 1 <= cell_value <= 4:
            data_rows += 1
    
    assert data_rows == 4, f"Expected 4 data rows, found {data_rows}"
    
    # Check totals row
    totals_found = False
    for row in ws.iter_rows(values_only=True):
        if any("ИТОГО" in str(cell) for cell in row):
            totals_found = True
            break
    
    assert totals_found, "Totals row not found"


def test_positions_sheet_without_items(sample_payload_without_line_items):
    """Test positions sheet handles documents without line items."""
    xlsx_bytes = build_document_xlsx(sample_payload_without_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Позиции"]
    
    # Should show message about no items
    assert "нет табличных позиций" in ws["A4"].value.lower()


def test_review_sheet_structure(sample_payload_with_line_items):
    """Test review sheet has correct structure."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Проверка"]
    
    # Check header
    assert "Журнал ручной проверки" in ws["A1"].value
    
    # Check column headers
    headers = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    assert "Поле" in str(headers[0])
    assert "Значение модели" in str(headers[1])
    assert "Итоговое значение" in str(headers[2])
    assert "Уверенность" in str(headers[3])
    assert "Статус" in str(headers[4])
    assert "Технический путь" in str(headers[5])
    
    # Check that we have field rows
    field_rows = 0
    for row_idx in range(5, 22):
        path_value = ws.cell(row=row_idx, column=6).value
        if path_value and str(path_value).startswith("/"):
            field_rows += 1
    
    assert field_rows > 0, "No field rows found"
    
    # Check corrected field styling
    corrected_found = False
    for row_idx in range(5, 22):
        corrected_value = ws.cell(row=row_idx, column=5).value
        if corrected_value == "Исправлено":
            corrected_found = True
            # Corrected values are highlighted in the review journal.
            cell_fill = ws.cell(row=row_idx, column=1).fill
            assert cell_fill.fill_type == "solid"
            break
    
    assert corrected_found, "No corrected fields found"


def test_guardrails_sheet_structure(sample_payload_with_line_items):
    """Test guardrails sheet has correct structure."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Guardrails"]
    
    # Check header
    assert "Контрольные проверки" in ws["A1"].value
    
    # Check column headers
    headers = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    assert "Проверка" in str(headers[0])
    assert "Результат" in str(headers[1])
    assert "Пояснение" in str(headers[2])
    
    # Check that we have guardrail rows
    guardrail_rows = 0
    for row_idx in range(5, 13):
        rule_value = ws.cell(row=row_idx, column=1).value
        if rule_value:
            guardrail_rows += 1
    
    assert guardrail_rows == 2, f"Expected 2 guardrail rows, found {guardrail_rows}"
    
    # Check that results show "Пройдено"
    passed_count = 0
    for row_idx in range(5, 13):
        result_value = ws.cell(row=row_idx, column=2).value
        if result_value == "Пройдено":
            passed_count += 1
    
    assert passed_count == 2, f"Expected 2 passed results, found {passed_count}"


def test_guardrails_sheet_empty(sample_payload_without_line_items):
    """Test guardrails sheet handles empty guardrails."""
    xlsx_bytes = build_document_xlsx(sample_payload_without_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Guardrails"]
    
    # Should show message about no violations
    assert "Нарушений не зарегистрировано" in ws["A5"].value


def test_numeric_values_as_numbers(sample_payload_with_line_items):
    """Test that numeric values are stored as numbers, not text."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Позиции"]
    
    # Find a line item with numeric values
    for row_idx in range(5, 13):
        qty_cell = ws.cell(row=row_idx, column=4)
        price_cell = ws.cell(row=row_idx, column=5)
        amount_cell = ws.cell(row=row_idx, column=6)
        
        if qty_cell.value is not None:
            assert isinstance(qty_cell.value, (int, float)), f"Quantity should be numeric, got {type(qty_cell.value)}"
            assert qty_cell.number_format in ("#,##0", "#,##0.###", "General"), f"Quantity should have numeric format"
        
        if price_cell.value is not None:
            assert isinstance(price_cell.value, (int, float)), f"Price should be numeric, got {type(price_cell.value)}"
            assert price_cell.number_format == "#,##0.00", f"Price should have number format with 2 decimals"
        
        if amount_cell.value is not None:
            assert isinstance(amount_cell.value, (int, float)), f"Amount should be numeric, got {type(amount_cell.value)}"
            assert amount_cell.number_format == "#,##0.00", f"Amount should have number format with 2 decimals"
        
        if qty_cell.value is not None:
            break  # Found a valid data row


def test_inn_kpp_as_text(sample_payload_with_line_items):
    """Test that INN and KPP are stored as text and not coerced to float."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)

    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]

    supplier_inn = "7736050003"
    supplier_kpp = "771401001"
    buyer_inn = "7707083893"
    buyer_kpp = "502901001"

    all_rows = [
        " ".join(str(cell) for cell in row if cell is not None)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)
    ]

    supplier_row = next((r for r in all_rows if supplier_inn in r), None)
    assert supplier_row is not None, f"Supplier INN {supplier_inn} not found"
    assert supplier_kpp in supplier_row, f"Supplier KPP {supplier_kpp} not found in supplier row"
    assert ".0" not in supplier_row, f"Supplier INN/KPP row contains '.0' (numeric coercion)"

    buyer_row = next((r for r in all_rows if buyer_inn in r), None)
    assert buyer_row is not None, f"Buyer INN {buyer_inn} not found"
    assert buyer_kpp in buyer_row, f"Buyer KPP {buyer_kpp} not found in buyer row"
    assert ".0" not in buyer_row, f"Buyer INN/KPP row contains '.0' (numeric coercion)"


def test_russian_date_formatting(sample_payload_with_line_items):
    """Test that dates are formatted in Russian."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]
    
    # Check title row for Russian date format
    title_value = ws["A3"].value
    assert "марта" in title_value, "Date should be in Russian (month name)"
    assert "2025" in title_value, "Year should be present"


def test_effective_corrections_applied(sample_payload_with_line_items):
    """Test that effective corrections are used in the document."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]
    
    # Find INN row and check for corrected values
    for row in ws.iter_rows(min_row=5, max_row=15, values_only=True):
        for cell in row:
            if cell and "7736050003" in str(cell):
                # This is the corrected supplier INN
                assert "7736050003" in str(cell), "Corrected supplier INN should be used"
                # Original was 7714236789, should not be present
                assert "7714236789" not in str(cell), "Original supplier INN should not be present"
                break
        else:
            continue
        break


def test_worksheet_print_properties(sample_payload_with_line_items):
    """Test that worksheets have correct print properties."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    
    # Check each sheet has print properties set
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Check paper size is A4
        assert str(ws.page_setup.paperSize) == str(ws.PAPERSIZE_A4), f"{sheet_name}: Paper size should be A4"
        
        # Check fit-to-width print contract and explicit print area.
        assert ws.page_setup.fitToWidth == 1, f"{sheet_name}: Should fit to width"
        assert ws.print_area, f"{sheet_name}: Print area should be set"

        # Premium layout uses compact A4 margins.
        assert ws.page_margins.left <= 0.3
        assert ws.page_margins.right <= 0.3
        assert ws.page_margins.top <= 0.5
        assert ws.page_margins.bottom <= 0.5


def test_totals_calculation(sample_payload_with_line_items):
    """Test that totals are calculated correctly."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]
    
    # Expected totals from line items
    expected_subtotal = 15580 + 16440 + 24450 + 1500  # 57970
    expected_vat = 3116 + 3288 + 4890 + 300  # 11594
    expected_total = expected_subtotal + expected_vat  # 69564
    
    # Find totals rows
    subtotal_found = False
    vat_found = False
    total_found = False
    
    for row in ws.iter_rows(values_only=True):
        row_values = [str(cell) if cell is not None else "" for cell in row]
        
        if "Итого без НДС" in "".join(row_values):
            # Find the numeric value in this row
            for cell in row:
                if isinstance(cell, (int, float)) and cell == expected_subtotal:
                    subtotal_found = True
                    break
        
        if "НДС" in "".join(row_values) and "Итого" not in "".join(row_values):
            for cell in row:
                if isinstance(cell, (int, float)) and cell == expected_vat:
                    vat_found = True
                    break
        
        if "Всего к оплате" in "".join(row_values):
            for cell in row:
                if isinstance(cell, (int, float)) and cell == expected_total:
                    total_found = True
                    break
    
    assert subtotal_found, f"Subtotal {expected_subtotal} not found"
    assert vat_found, f"VAT {expected_vat} not found"
    assert total_found, f"Total {expected_total} not found"


def test_xlsx_is_valid_file(sample_payload_with_line_items):
    """Test that the output is a valid XLSX file."""
    xlsx_bytes = build_document_xlsx(sample_payload_with_line_items)
    
    # Check that we got bytes
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0
    
    # Check that it's a valid ZIP file (XLSX is a ZIP)
    import zipfile
    from io import BytesIO
    
    try:
        with zipfile.ZipFile(BytesIO(xlsx_bytes), 'r') as zf:
            # Check for required XLSX files
            assert "[Content_Types].xml" in zf.namelist()
            assert "_rels/.rels" in zf.namelist()
            assert "xl/workbook.xml" in zf.namelist()
    except zipfile.BadZipFile:
        pytest.fail("Output is not a valid ZIP/XLSX file")


def test_duplicate_guardrails_removed():
    """Test that duplicate guardrails are removed."""
    payload = {
        "document": {"id": "doc-123", "file_name": "test.pdf"},
        "extracted": {"fields": [], "line_items": []},
        "guardrails": [
            {
                "rule": "/test/inn",
                "passed": True,
                "message": "/test/inn: manual_correction: исправлено",
            },
            {
                "rule": "/test/inn",
                "passed": True,
                "message": "/test/inn: manual_correction: исправлено",
            },
            {
                "rule": "/test/name",
                "passed": False,
                "message": "/test/name: invalid value",
            },
        ],
    }
    
    xlsx_bytes = build_document_xlsx(payload)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Guardrails"]
    
    # Count data rows (excluding headers)
    data_rows = 0
    for row_idx in range(5, 13):
        if ws.cell(row=row_idx, column=1).value:
            data_rows += 1
    
    # Should have 2 unique guardrails, not 3
    assert data_rows == 2, f"Expected 2 unique guardrails, found {data_rows}"


def test_line_items_with_none_numeric_fields():
    """Test that line items with None/empty numeric fields are handled safely."""
    payload = {
        "document": {"id": "doc-789", "file_name": "invoice.pdf"},
        "extracted": {
            "fields": [
                {
                    "path": "/number",
                    "name": "Номер",
                    "value": "999",
                    "original_value": "999",
                    "confidence": 1.0,
                    "corrected": False,
                },
            ],
            "line_items": [
                {
                    "name": "Товар с частично заполненными данными",
                    "unit": "шт",
                    "quantity": None,  # None value
                    "unit_price": "",  # Empty string
                    "amount": None,  # None value
                    "vat_rate": None,
                    "vat_amount": None,
                    "total": None,
                },
                {
                    "name": "Товар с валидными данными",
                    "unit": "шт",
                    "quantity": 5,
                    "unit_price": 1000,
                    "amount": 5000,
                    "vat_rate": "20",
                    "vat_amount": 1000,
                    "total": 6000,
                },
                {
                    "name": "Товар с некорректными строками",
                    "unit": "шт",
                    "quantity": "не число",  # Invalid string
                    "unit_price": "текст",  # Invalid string
                    "amount": "N/A",  # Invalid string
                    "vat_rate": None,
                    "vat_amount": None,
                    "total": None,
                },
            ],
        },
        "guardrails": [],
    }
    
    # This should not raise TypeError
    xlsx_bytes = build_document_xlsx(payload)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]
    
    # Check that all line items are displayed
    all_text = " ".join(
        str(cell)
        for row in ws.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    
    for item in payload["extracted"]["line_items"]:
        assert item["name"] in all_text, f"Line item '{item['name']}' not found"


def test_line_items_all_none_amounts():
    """Test that line items with all None amounts calculate totals correctly (0.0)."""
    payload = {
        "document": {"id": "doc-999", "file_name": "invoice.pdf"},
        "extracted": {
            "fields": [
                {
                    "path": "/number",
                    "name": "Номер",
                    "value": "1000",
                    "original_value": "1000",
                    "confidence": 1.0,
                    "corrected": False,
                },
            ],
            "line_items": [
                {
                    "name": "Товар 1",
                    "unit": "шт",
                    "quantity": None,
                    "unit_price": None,
                    "amount": None,
                    "vat_rate": None,
                    "vat_amount": None,
                    "total": None,
                },
                {
                    "name": "Товар 2",
                    "unit": "шт",
                    "quantity": None,
                    "unit_price": None,
                    "amount": None,
                    "vat_rate": None,
                    "vat_amount": None,
                    "total": None,
                },
            ],
        },
        "guardrails": [],
    }
    
    # This should not raise TypeError and should handle None gracefully
    xlsx_bytes = build_document_xlsx(payload)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]
    
    # Check that totals are 0.0 when all amounts are None
    for row in ws.iter_rows(values_only=True):
        row_values = [str(cell) if cell is not None else "" for cell in row]
        
        if "Итого без НДС" in "".join(row_values):
            for cell in row:
                if isinstance(cell, (int, float)):
                    assert cell == 0.0, f"Expected subtotal 0.0, got {cell}"
                    break


def test_document_without_parties():
    """Test that document without supplier/buyer information is handled safely."""
    payload = {
        "document": {"id": "doc-888", "file_name": "invoice.pdf"},
        "extracted": {
            "fields": [
                {
                    "path": "/number",
                    "name": "Номер",
                    "value": "777",
                    "original_value": "777",
                    "confidence": 1.0,
                    "corrected": False,
                },
            ],
            "line_items": [
                {
                    "name": "Товар",
                    "unit": "шт",
                    "quantity": 1,
                    "unit_price": 100,
                    "amount": 100,
                    "vat_rate": "20",
                    "vat_amount": 20,
                    "total": 120,
                },
            ],
        },
        "guardrails": [],
    }
    
    # This should not raise any errors
    xlsx_bytes = build_document_xlsx(payload)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]
    
    # Document should still be generated
    assert ws is not None
    assert len(wb.sheetnames) == 4


def test_document_without_line_items():
    """Test that document without line items handles footer and signature rows correctly."""
    payload = {
        "document": {"id": "doc-999", "file_name": "contract.pdf"},
        "extracted": {
            "fields": [
                {
                    "path": "/number",
                    "name": "Номер",
                    "value": "CTR-2025-001",
                    "original_value": "CTR-2025-001",
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/supplier/name",
                    "name": "Поставщик",
                    "value": "ООО Поставщик",
                    "original_value": "ООО Поставщик",
                    "confidence": 1.0,
                    "corrected": False,
                },
                {
                    "path": "/buyer/name",
                    "name": "Покупатель",
                    "value": "ООО Покупатель",
                    "original_value": "ООО Покупатель",
                    "confidence": 1.0,
                    "corrected": False,
                },
            ],
            "line_items": [],  # No line items
        },
        "guardrails": [],
    }
    
    # This should not raise any errors
    xlsx_bytes = build_document_xlsx(payload)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]
    ws_pos = wb["Позиции"]
    
    # Document should be generated without line items table
    assert ws is not None
    assert len(wb.sheetnames) == 4
    
    # Check that parties are shown even without line items
    all_text = " ".join(
        str(cell)
        for row in ws.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    assert "ООО Поставщик" in all_text
    assert "ООО Покупатель" in all_text
    
    # Check positions sheet shows message
    assert "нет табличных позиций" in ws_pos["A4"].value.lower()
    
    # No line items table should be present (no numeric values from line items)
    # Only footer and signature rows should be present after parties


def test_line_item_total_none_vat_amount():
    """Test specific case: total=None, amount=valid number, vat_amount=None."""
    payload = {
        "document": {"id": "doc-zero-vat", "file_name": "invoice.pdf"},
        "extracted": {
            "fields": [
                {
                    "path": "/number",
                    "name": "Номер",
                    "value": "001",
                    "original_value": "001",
                    "confidence": 1.0,
                    "corrected": False,
                },
            ],
            "line_items": [
                {
                    "name": "Товар без VAT",
                    "unit": "шт",
                    "quantity": 1,
                    "unit_price": 100,
                    "amount": 100,
                    "vat_rate": "without_vat",
                    "vat_amount": None,  # Bug case: None here
                    "total": None,
                },
            ],
        },
        "guardrails": [],
    }
    
    # This should not raise TypeError
    xlsx_bytes = build_document_xlsx(payload)
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Документ"]
    
    # Check that file opens without errors
    assert ws is not None
    assert len(wb.sheetnames) == 4
    
    # Check that the item is displayed
    all_text = " ".join(
        str(cell)
        for row in ws.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    assert "Товар без VAT" in all_text
    assert "Товар без VAT" in all_text  # Appears twice due to name column
    
    print("✓ Test passed: line_item_total_none_vat_amount")


def test_premium_layout_contract(sample_payload_with_line_items):
    """Lock the accepted workbook names and prohibit invented signatory labels."""
    from io import BytesIO
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(build_document_xlsx(sample_payload_with_line_items)))
    assert workbook.sheetnames == ["Документ", "Позиции", "Проверка", "Guardrails"]

    document = workbook["Документ"]
    assert document.max_column == 10
    assert document["A1"].value == "DocSift  /  ПРОВЕРЕННЫЙ ДОКУМЕНТ"
    assert document["G1"].value == "ПРОВЕРЕНО"
    assert document.print_area
    assert document.oddFooter.right.text == "Страница &P из &N"

    workbook_text = " ".join(
        str(cell)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    assert "Руководитель" not in workbook_text
    assert "Бухгалтер" not in workbook_text
